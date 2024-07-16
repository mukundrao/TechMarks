import openpyxl
import math
from openpyxl.styles import PatternFill, Font

def process(init_file, fin_file, sheet_name):                   #this function processes 'input_file' and returns a list of dictionaries, each for a student
    wb = openpyxl.load_workbook(init_file)                      #opening the 'input_file' excel file for processing
    sheet = wb[sheet_name]                                      #accessing the required sheet to be processed in 'input_file'
    data = []
    d = {}
    font_style = Font(bold=True)

    for row in range(2, sheet.max_row + 1):                     #for loop to process data row wise
        name = d['name'] = sheet.cell(row, 2).value

        try:
            cie1 = d['cie1'] = sheet.cell(row, 6).value = sheet.cell(row, 5).value * 0.5    #accessing the value of the cell in the 5th column of that row and calculating the value for the cell in the 6th column of the same row
            sheet.cell(row, 6).font = font_style                                            #changing the font style of text in the cell present in the 6th column of a row
        except TypeError:                                                                   #exception to handle no-float values in the cell in the 6th column of a row
            cie1 = d['cie1'] = sheet.cell(row, 6).value = 0 * 0.5
            sheet.cell(row, 6).font = font_style

        try:
            cie2 = d['cie2'] = sheet.cell(row, 8).value = sheet.cell(row, 7).value * 0.5
            sheet.cell(row, 8).font = font_style
        except TypeError:
            cie2 = d['cie2'] = sheet.cell(row, 8).value = 0 * 0.5
            sheet.cell(row, 8).font = font_style

        try:
            cie3 = d['cie3'] = sheet.cell(row, 10).value = sheet.cell(row, 9).value * 0.5
            sheet.cell(row, 10).font = font_style
        except TypeError:
            cie3 = d['cie3'] = sheet.cell(row, 10).value = 0 * 0.5
            sheet.cell(row, 10).font = font_style

        try:
            aat1 = d['aat1'] = sheet.cell(row, 12).value = sheet.cell(row, 11).value * 0.5
            sheet.cell(row, 12).font = font_style
        except TypeError:
            aat1 = d['aat1'] = sheet.cell(row, 12).value = 0 * 0.5
            sheet.cell(row, 12).font = font_style

        try:
            aat2 = d['aat2'] = sheet.cell(row, 14).value = sheet.cell(row, 13).value * 0.5
            sheet.cell(row, 14).font = font_style
        except TypeError:
            aat2 = d['aat2'] = sheet.cell(row, 14).value = 0 * 0.5
            sheet.cell(row, 14).font = font_style

        cie = d['cie'] = sheet.cell(row, 15).value = math.ceil(cie1 + cie2 + cie3 - min(cie1, cie2, cie3))
        sheet.cell(row, 15).fill = PatternFill(patternType='solid', fgColor='dda8ed')      #filling the cell in the 15th column of a row with color having a hex code 'dda8ed'
        sheet.cell(row, 15).font = font_style

        aat = d['aat'] = sheet.cell(row, 16).value = math.ceil(aat1 + aat2)
        sheet.cell(row, 16).fill = PatternFill(patternType='solid', fgColor='dda8ed')
        sheet.cell(row, 16).font = font_style

        internals = d['internals'] = sheet.cell(row, 17).value = cie + aat
        sheet.cell(row, 17).fill = PatternFill(patternType='solid', fgColor='8ee2e8')
        sheet.cell(row, 17).font = font_style

        attendance = d['attendance'] = sheet.cell(row, 18).value
        sheet.cell(row, 19).font = font_style

        #below is the condition to check the eligibility of a student for writing SEE and performing further conditional styling:
        if ((cie1 >= 8 and cie2 >= 8) or (cie1 >= 8 and cie3 >= 8) or (cie2 >= 8 and cie3 >= 8)) and aat1 >= 2 and aat2 >= 2 and attendance >= 85:
            eligibility = d['eligibility'] = sheet.cell(row, 19).value = 'Eligible'

        else:
            eligibility = d['eligibility'] = sheet.cell(row, 19).value = 'Not Eligible'
            sheet.cell(row, 19).fill = PatternFill(patternType='solid', fgColor='f06969')
            sheet.cell(row, 2).fill = PatternFill(patternType='solid', fgColor='f06969')
            sheet.cell(row, 3).fill = PatternFill(patternType='solid', fgColor='f06969')

            if cie1 < 8 and cie2 < 8:
                sheet.cell(row, 6).fill = PatternFill(patternType='solid', fgColor='f06969')
                sheet.cell(row, 8).fill = PatternFill(patternType='solid', fgColor='f06969')

            if cie2 < 8 and cie3 < 8:
                sheet.cell(row, 8).fill = PatternFill(patternType='solid', fgColor='f06969')
                sheet.cell(row, 10).fill = PatternFill(patternType='solid', fgColor='f06969')

            if cie1 < 8 and cie3 < 8:
                sheet.cell(row, 6).fill = PatternFill(patternType='solid', fgColor='f06969')
                sheet.cell(row, 10).fill = PatternFill(patternType='solid', fgColor='f06969')

            if aat1 < 2:
                sheet.cell(row, 12).fill = PatternFill(patternType='solid', fgColor='f06969')

            if aat2 < 2:
                sheet.cell(row, 14).fill = PatternFill(patternType='solid', fgColor='f06969')

            if attendance < 85:
                sheet.cell(row, 18).fill = PatternFill(patternType='solid', fgColor='f06969')

        email = d['email'] = sheet.cell(row, 4).value

        data.append(d.copy())                                #storing the dictionary of each student in a list which will be passed to function call in 'main.py'

    wb.save(fin_file)                                        #saving the processed excel file to the project folder
    return data                                              # returning the list data that contains dictionaries of all students which in turn contain all student data
